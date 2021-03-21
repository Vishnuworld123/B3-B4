# guido van rossum - 
FILE_PATH = "F:\\Class\\B3-B4\\Excel\\first_file.xlsx"

"""
virtual env ---

Django --- 

software comp - 2 projects   ---
1 - project user - django 2.0   --
2 - project     - django 3.0  # addon



install django 2.0   - proect 1 --- project run   - system env - 
virtual env env1 -1    - django 2.0
virtual envenv2 - 2   - django 3.0

env - 1 --- phycsic
env 2 -   chemistry
env - 3  - maths


common hall - 
college - leacture - 1 hall - physics 
 2 -- chemistry - 


"""


from openpyxl import Workbook, load_workbook
import time  
import openpyxl

# wb = Workbook()  
# sheet = wb.active  

# sheet['A1'] = 87  
# sheet['A2'] = "Devansh"  
# sheet['A3'] = 41.80  
# sheet['A4'] = 10 

# now = time.strftime("%x")  
# print(now)
# sheet['A5'] = now  

# sheet.cell(1,1).value = 'abc'
# sheet.cell(1,2).value = 'xyz'
# sheet.cell(1,3).value = 4545
# sheet.cell(1,4).value = 'asbdja'
# sheet.cell(1,5).value = now

# wb.save("F:\\Class\\B3-B4\\Excel\\first_file.xlsx") 

# wb = load_workbook("F:\\Class\\B3-B4\\Excel\\first_file.xlsx")  

# sheet = wb.active  
# 1st way to put the values  in a cell
# sheet['A1'] = 'Devansh Sharma'  
  

# 2nd way to put the values  in a cell
# sheet.cell(row=2, column=2).value = 5  

# wb.save("F:\\Class\\B3-B4\\Excel\\first_file.xlsx") 

#####
  
# wb = Workbook()  
# sheet = wb.active  
  
# sheet['A1'] = "id"
# sheet['B1'] = "data"

# sheet.cell(row=1, column=3).value = 'value'

# data = (  
#     (11, 48, 50),  
#     (81, 30, 82),  
#     (20, 51, 72),  
#     (21, 14, 60),  
#     (28, 41, 49),  
#     (74, 65, 53),  
#     ("Peter", 'Andrew',45.63)  
# )  
  
# for i in data:  
#     sheet.append(i) 

# wb.save("F:\\Class\\B3-B4\\Excel\\first_file.xlsx") 


###


# wb = load_workbook(FILE_PATH)  

# sheet = wb.active  

# x1 = sheet['A1'].value  
# x2 = sheet['A2'].value  
#using cell() function  
# x3 = sheet.cell(row=3, column=1).value  
  
# print("The first cell value:",x1)  
# print("The second cell value:",x2)  
# print("The third cell value:",x3) 


#########


# wb = load_workbook(FILE_PATH)  
  
# sheet = wb.active  


# cells = sheet['A1':'C7']  
# print(cells)

# for i1, i2, i3 in cells:  
#     print(i1.value, i2.value, i3.value)


##########

# from openpyxl import Workbook  
  
# wb = Workbook()  
# sheet = wb.active  
  
# rows = (  
#     (90, 46, 48, 44),  
#     (81, 30, 32, 16),  
#     (23, 95, 87,27),  
#     (65, 12, 89, 53),  
#     (42, 81, 40, 44),  
#     (34, 51, 76, 42)  
# )  
  
# for row in rows:  
#     sheet.append(row)  

# print(sheet.max_row)
# print(sheet.max_column)

# for row in sheet.iter_rows(min_row=sheet.min_row, min_col=sheet.min_column, max_row=sheet.max_row, max_col=sheet.max_column):  
    # print(row)  # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>, <Cell 'Sheet'.D1>)
    # l = []
    # for cell in row: 
        # print(cell.value, end=' ')  
        # l.append(cell.value)
    # print() 
    # print(l)


# wb.save(FILE_PATH) 


#######################################


# wb = openpyxl.load_workbook(FILE_PATH)
# sheet = wb.active 


# for i in range(1, 11):  # 2
#     for j in range(1, 11):  # 1 
#         sheet.cell(row=i,column=j).value = i

# wb.save(FILE_PATH)



