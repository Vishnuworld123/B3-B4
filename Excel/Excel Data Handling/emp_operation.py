from emp_classes import Employee, Address
from utility import generate_emps
from emp_abstraction import EmpAbstraction
import openpyxl
from openpyxl.styles import Alignment  

FILE_PATH = r"F:\Class\B3-B4\Excel\Excel Data Handling\emp.xlsx"

def get_sheet_and_workbook():
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("Emp Data")
    wb.remove(wb["Sheet"])
    wb.save(FILE_PATH)
    return wb, sheet

def fetch_sheet_and_workbook():
    wb = openpyxl.load_workbook(FILE_PATH)
    sheet = wb['Emp Data']
    return wb, sheet


class EmployeeImplementation(EmpAbstraction):
    def write_headers(self):
        """to write headers into excel"""
        header_line = "ID Name Age Salary Address"
        header_list = header_line.split()
        wb, sheet = get_sheet_and_workbook()  #tuple
        c = 1
        for i in header_list:
            sheet.cell(1, c).value = i.upper()  # 1, 1  = ID, Name
            c +=1 
        sheet.merge_cells("E1:G1")

        for i in range(1, sheet.max_column):
            cell = sheet.cell(row=1, column=i)  
            cell.alignment = Alignment(horizontal='center', vertical='center') 
        address_header = "Pincode Area City"

        # write sub address params in second row
        count = 5
        for i in address_header.split():
            sheet.cell(row=2, column=count).value = i.upper()
            count += 1
        wb.save(FILE_PATH)
        return "Headers created successfully..!"

    def write_data(self, no):
        wb, sheet = fetch_sheet_and_workbook()
        c = 2
        emps = generate_emps(no)  # list of emp dict
        for e in emps:
            sheet.cell(row=1+c, column=1).value = e.EmpID
            sheet.cell(row=1+c, column=2).value = e.EmpName
            sheet.cell(row=1+c, column=3).value = e.EmpAge
            sheet.cell(row=1+c, column=4).value = e.EmpSalary
            sheet.cell(row=1+c, column=5).value = e.EmpAddress.Pincode
            sheet.cell(row=1+c, column=6).value = e.EmpAddress.Area
            sheet.cell(row=1+c, column=7).value = e.EmpAddress.City
            c += 1
        wb.save(FILE_PATH)
        return "Data inserted successfully..!"

    def read_data(self):
        wb, sheet = fetch_sheet_and_workbook()
        final_emp_list = []
        for i in range(3, sheet.max_row+1): # 3, 52
            empid = sheet.cell(row=i, column=1).value
            empname = sheet.cell(row=i, column=2).value
            empage = sheet.cell(row=i, column=3).value
            empsal = sheet.cell(row=i, column=4).value
            adrpin = sheet.cell(row=i, column=5).value
            adrare = sheet.cell(row=i, column=6).value
            adrcit = sheet.cell(row=i, column=7).value
            emp = Employee(eid=empid, ename=empname, eage=empage, esalary=empsal, eadr=Address(pin=adrpin, area=adrare, city=adrcit))
            final_emp_list.append(emp)

        return final_emp_list

# emp_imp = EmployeeImplementation()
# emp_imp.write_headers()
# emp_imp.write_data(no=50)
# print(emp_imp.read_data())

# [{'EmpID': 101, 'EmpName': 'Ydmk', 'EmpAge': 35, 'EmpSalary': 40055, 'EmpAddress': [{'Pincode': 429119, 'Area': 'Karve Nagar', 'City': 'Pune'},
# {'Pincode': 429119, 'Area': 'Karve Nagar', 'City': 'Pune'}, {'Pincode': 429119, 'Area': 'Karve Nagar', 'City': 'Pune'}}],]


# task - salary - average salary for all emps -
# def get_avg_salary

#  lambda filter, salary 50k +, show

# city merge, alignment centre, using one for loop
# multiple address - 

# MySQL -- database install -- 





